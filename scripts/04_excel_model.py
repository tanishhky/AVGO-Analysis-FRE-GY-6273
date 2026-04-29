"""
================================================================================
04_excel_model.py
Builds a multi-sheet Excel financial model for AVGO with proper formulas,
NYU Violet color scheme, and standard financial-modeling color coding.
================================================================================
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os

# --- Portable path resolution (relative to script location) ---
from pathlib import Path
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

OUT_DIR = str(PROJECT_ROOT / 'output')
DATA_DIR = str(PROJECT_ROOT / 'data')
os.makedirs(OUT_DIR, exist_ok=True)
OUTPUT_PATH = f'{OUT_DIR}/AVGO_Financial_Model.xlsx'

# NYU Violet styling
NYU_VIOLET = '57068C'
NYU_LIGHT  = '8900E1'
NYU_DARK   = '330050'
NYU_LGREY  = 'EFEFEF'
ACCENT_GOLD = 'C99700'

# Industry-standard financial-model color coding
BLUE   = '0000FF'   # hardcoded inputs
BLACK  = '000000'   # formulas
GREEN  = '008000'   # cross-sheet links
YELLOW = 'FFFF00'   # key assumptions

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK = Side(style='medium', color=NYU_VIOLET)
BORDER_HEADER = Border(bottom=THICK)

wb = Workbook()
wb.remove(wb.active)


def style_header(cell, fill=NYU_VIOLET, color='FFFFFF', size=11):
    cell.font = Font(name='Arial', bold=True, color=color, size=size)
    cell.fill = PatternFill('solid', start_color=fill)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def style_input(cell, fmt='#,##0'):
    cell.font = Font(name='Arial', color=BLUE)
    cell.number_format = fmt


def style_formula(cell, fmt='#,##0'):
    cell.font = Font(name='Arial', color=BLACK)
    cell.number_format = fmt


def style_link(cell, fmt='#,##0'):
    cell.font = Font(name='Arial', color=GREEN)
    cell.number_format = fmt


def style_label(cell, bold=False):
    cell.font = Font(name='Arial', bold=bold, color=BLACK)


# ============================================================
# SHEET 1: Cover
# ============================================================
ws = wb.create_sheet('1. Cover')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 22
for col in 'DEFGH':
    ws.column_dimensions[col].width = 18

ws['B2'] = 'WASHINGTON SQUARE RESEARCH PARTNERS'
ws['B2'].font = Font(name='Arial', bold=True, size=18, color=NYU_VIOLET)
ws['B3'] = 'NYU MSFE Equity Research | Equity Research Report — AVGO'
ws['B3'].font = Font(name='Arial', italic=True, size=12, color=NYU_DARK)
ws['B4'] = 'Authors: Tanishk Yadav | Martin Baretto | Ramana Sriram'
ws['B4'].font = Font(name='Arial', size=10, color='666666')
ws['B5'] = 'Report date: 25 April 2026'
ws['B5'].font = Font(name='Arial', size=10, color='666666')

ws['B7'] = 'Recommendation'
ws['B7'].font = Font(name='Arial', bold=True, size=12)
ws['C7'] = 'BUY'
ws['C7'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
ws['C7'].fill = PatternFill('solid', start_color=ACCENT_GOLD)
ws['C7'].alignment = Alignment(horizontal='center')

ws['B8'] = 'Spot price (USD)'
ws['C8'] = 422.76
style_input(ws['C8'], fmt='$#,##0.00')

ws['B9'] = '12-month price target (USD)'
ws['C9'] = 475.00
style_input(ws['C9'], fmt='$#,##0.00')

ws['B10'] = 'Implied upside'
ws['C10'] = '=C9/C8-1'
style_formula(ws['C10'], fmt='0.0%')

ws['B11'] = 'Market cap ($mn)'
ws['C11'] = 2001630
style_input(ws['C11'])

ws['B12'] = 'Diluted shares (mn)'
ws['C12'] = 4969
style_input(ws['C12'])

ws['B13'] = '52-week range'
ws['C13'] = '$138.10 – $422.76'
ws['C13'].alignment = Alignment(horizontal='center')

# Sheet directory
ws['B15'] = 'Workbook Contents'
ws['B15'].font = Font(name='Arial', bold=True, size=12, color=NYU_VIOLET)
sheets = [
    ('1. Cover', 'This page'),
    ('2. Assumptions', 'Key drivers, growth rates, margins'),
    ('3. Income Statement', 'FY23A – FY30E with formulas'),
    ('4. DCF', 'FCFF + Gordon terminal value'),
    ('5. WACC', 'Cost of capital build'),
    ('6. Sensitivity', 'WACC × terminal-g grid'),
    ('7. Comps', 'Peer trading multiples'),
    ('8. SOTP', 'Sum-of-the-parts by segment'),
    ('9. Monte Carlo', 'Stochastic distribution summary'),
    ('10. Football Field', 'Synthesis of valuation methods'),
]
for i, (name, desc) in enumerate(sheets, start=16):
    ws[f'B{i}'] = name
    ws[f'C{i}'] = desc
    ws[f'B{i}'].font = Font(name='Arial', bold=True, color=NYU_VIOLET)
    ws[f'C{i}'].font = Font(name='Arial', color='333333')

ws['B28'] = 'Color coding (industry standard)'
ws['B28'].font = Font(name='Arial', bold=True, color=NYU_DARK)
ws['B29'] = '  Blue text';  ws['B29'].font = Font(name='Arial', color=BLUE);  ws['C29'] = 'Hardcoded inputs'
ws['B30'] = '  Black text'; ws['B30'].font = Font(name='Arial', color=BLACK); ws['C30'] = 'Formulas / calculations'
ws['B31'] = '  Green text'; ws['B31'].font = Font(name='Arial', color=GREEN); ws['C31'] = 'Cross-sheet links'
ws['B32'] = '  Yellow fill'; ws['B32'].fill = PatternFill('solid', start_color=YELLOW); ws['C32'] = 'Key assumptions'

# ============================================================
# SHEET 2: Assumptions
# ============================================================
ws = wb.create_sheet('2. Assumptions')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 36
for col in 'CDEFGHIJK':
    ws.column_dimensions[col].width = 13

ws['B2'] = 'OPERATING ASSUMPTIONS'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

# Year headers
years = ['FY23A', 'FY24A', 'FY25A', 'FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
for i, y in enumerate(years):
    cell = ws.cell(row=4, column=3 + i, value=y)
    style_header(cell)

ws['B5'] = 'Revenue ($mn)'
ws['B5'].font = Font(name='Arial', bold=True, color=NYU_DARK)

ws['B6'] = '  AI Semiconductor'
ws.cell(row=6, column=3, value=4000);    style_input(ws.cell(row=6, column=3))
ws.cell(row=6, column=4, value=12200);   style_input(ws.cell(row=6, column=4))
ws.cell(row=6, column=5, value=24500);   style_input(ws.cell(row=6, column=5))
for i, v in enumerate([60000, 115000, 165000, 205000, 235000]):
    c = ws.cell(row=6, column=6 + i, value=v); style_input(c)
    c.fill = PatternFill('solid', start_color=YELLOW)

ws['B7'] = '  Non-AI Semiconductor'
ws.cell(row=7, column=3, value=21800);   style_input(ws.cell(row=7, column=3))
ws.cell(row=7, column=4, value=18500);   style_input(ws.cell(row=7, column=4))
ws.cell(row=7, column=5, value=15500);   style_input(ws.cell(row=7, column=5))
for i, v in enumerate([16500, 17000, 17750, 18500, 19250]):
    c = ws.cell(row=7, column=6 + i, value=v); style_input(c)
    c.fill = PatternFill('solid', start_color=YELLOW)

ws['B8'] = '  Infrastructure Software'
ws.cell(row=8, column=3, value=10000);   style_input(ws.cell(row=8, column=3))
ws.cell(row=8, column=4, value=20800);   style_input(ws.cell(row=8, column=4))
ws.cell(row=8, column=5, value=23900);   style_input(ws.cell(row=8, column=5))
for i, v in enumerate([28500, 31000, 33500, 35500, 37500]):
    c = ws.cell(row=8, column=6 + i, value=v); style_input(c)
    c.fill = PatternFill('solid', start_color=YELLOW)

ws['B9'] = 'Total Revenue (formula)'
ws['B9'].font = Font(name='Arial', bold=True)
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=9, column=3 + i, value=f'=SUM({col}6:{col}8)')
    style_formula(ws.cell(row=9, column=3 + i))
    ws.cell(row=9, column=3 + i).font = Font(name='Arial', bold=True)

ws['B10'] = 'Revenue growth YoY'
for i in range(1, 8):
    cur_col = get_column_letter(3 + i)
    prev_col = get_column_letter(2 + i)
    ws.cell(row=10, column=3 + i, value=f'={cur_col}9/{prev_col}9-1')
    style_formula(ws.cell(row=10, column=3 + i), fmt='0.0%')

# Margin assumptions block
ws['B12'] = 'MARGIN & CAPITAL ASSUMPTIONS'
ws['B12'].font = Font(name='Arial', bold=True, size=12, color=NYU_DARK)

margin_rows = [
    ('Gross margin',    [0.696, 0.626, 0.745, 0.770, 0.775, 0.780, 0.780, 0.780]),
    ('Opex % of revenue',[0.085, 0.090, 0.085, 0.085, 0.075, 0.072, 0.070, 0.069]),
    ('D&A % of revenue', [0.030, 0.025, 0.020, 0.012, 0.010, 0.009, 0.009, 0.008]),
    ('Capex % of revenue',[0.018, 0.020, 0.012, 0.012, 0.010, 0.010, 0.010, 0.010]),
    ('NWC change % of revenue',[0.008, 0.010, 0.009, 0.010, 0.012, 0.010, 0.008, 0.008]),
    ('Effective tax rate', [0.18, 0.18, 0.17, 0.18, 0.18, 0.18, 0.18, 0.18]),
]
for r, (label, vals) in enumerate(margin_rows, start=13):
    ws.cell(row=r, column=2, value=label)
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=3 + i, value=v)
        style_input(c, fmt='0.0%')
        if r == 13 and 6 <= 3 + i <= 10:   # gross margin, projection
            c.fill = PatternFill('solid', start_color=YELLOW)


# ============================================================
# SHEET 3: Income Statement
# ============================================================
ws = wb.create_sheet('3. Income Statement')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 32
for col in 'CDEFGHIJK':
    ws.column_dimensions[col].width = 13

ws['B2'] = 'INCOME STATEMENT — FY23A – FY30E ($ in millions)'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

for i, y in enumerate(years):
    style_header(ws.cell(row=4, column=3 + i, value=y))

# AI Semi (link from Assumptions)
ws['B6'] = 'AI Semiconductor revenue'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=6, column=3 + i, value=f"='2. Assumptions'!{col}6")
    style_link(ws.cell(row=6, column=3 + i))

ws['B7'] = 'Non-AI Semi revenue'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=7, column=3 + i, value=f"='2. Assumptions'!{col}7")
    style_link(ws.cell(row=7, column=3 + i))

ws['B8'] = 'Infrastructure Software revenue'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=8, column=3 + i, value=f"='2. Assumptions'!{col}8")
    style_link(ws.cell(row=8, column=3 + i))

ws['B9'] = 'Total Revenue'
ws['B9'].font = Font(name='Arial', bold=True)
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=9, column=3 + i, value=f'=SUM({col}6:{col}8)')
    style_formula(c)
    c.font = Font(name='Arial', bold=True)
    c.fill = PatternFill('solid', start_color=NYU_LGREY)

ws['B10'] = 'Revenue growth YoY'
for i in range(1, 8):
    cur = get_column_letter(3 + i); prev = get_column_letter(2 + i)
    c = ws.cell(row=10, column=3 + i, value=f'={cur}9/{prev}9-1')
    style_formula(c, fmt='0.0%')

# Gross Profit
ws['B12'] = 'Gross margin %'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=12, column=3 + i, value=f"='2. Assumptions'!{col}13")
    style_link(ws.cell(row=12, column=3 + i), fmt='0.0%')

ws['B13'] = 'Gross profit'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=13, column=3 + i, value=f'={col}9*{col}12')
    style_formula(c)

# Opex & EBIT
ws['B14'] = 'Opex % of revenue'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=14, column=3 + i, value=f"='2. Assumptions'!{col}14")
    style_link(ws.cell(row=14, column=3 + i), fmt='0.0%')

ws['B15'] = 'Opex'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=15, column=3 + i, value=f'={col}9*{col}14')
    style_formula(c)

ws['B16'] = 'EBIT'
ws['B16'].font = Font(name='Arial', bold=True)
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=16, column=3 + i, value=f'={col}13-{col}15')
    style_formula(c); c.font = Font(name='Arial', bold=True)

ws['B17'] = 'EBIT margin %'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=17, column=3 + i, value=f'={col}16/{col}9')
    style_formula(c, fmt='0.0%')

# D&A and EBITDA
ws['B19'] = 'D&A % of revenue'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=19, column=3 + i, value=f"='2. Assumptions'!{col}15")
    style_link(ws.cell(row=19, column=3 + i), fmt='0.0%')

ws['B20'] = 'D&A'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=20, column=3 + i, value=f'={col}9*{col}19')
    style_formula(c)

ws['B21'] = 'EBITDA'
ws['B21'].font = Font(name='Arial', bold=True)
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=21, column=3 + i, value=f'={col}16+{col}20')
    style_formula(c); c.font = Font(name='Arial', bold=True)
    c.fill = PatternFill('solid', start_color=NYU_LGREY)

ws['B22'] = 'EBITDA margin %'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=22, column=3 + i, value=f'={col}21/{col}9')
    style_formula(c, fmt='0.0%')

# Tax & NOPAT
ws['B24'] = 'Effective tax rate'
for i in range(8):
    col = get_column_letter(3 + i)
    ws.cell(row=24, column=3 + i, value=f"='2. Assumptions'!{col}18")
    style_link(ws.cell(row=24, column=3 + i), fmt='0.0%')

ws['B25'] = 'Tax on EBIT'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=25, column=3 + i, value=f'={col}16*{col}24')
    style_formula(c)

ws['B26'] = 'NOPAT'
ws['B26'].font = Font(name='Arial', bold=True)
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=26, column=3 + i, value=f'={col}16-{col}25')
    style_formula(c); c.font = Font(name='Arial', bold=True)

# FCFF Build
ws['B28'] = 'FCFF BUILD'
ws['B28'].font = Font(name='Arial', bold=True, size=11, color=NYU_VIOLET)

ws['B29'] = 'Capex'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=29, column=3 + i,
                value=f"=-{col}9*'2. Assumptions'!{col}16")
    style_formula(c)

ws['B30'] = 'NWC change'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=30, column=3 + i,
                value=f"=-{col}9*'2. Assumptions'!{col}17")
    style_formula(c)

ws['B31'] = 'FCFF'
ws['B31'].font = Font(name='Arial', bold=True)
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=31, column=3 + i, value=f'={col}26+{col}20+{col}29+{col}30')
    style_formula(c); c.font = Font(name='Arial', bold=True)
    c.fill = PatternFill('solid', start_color=NYU_LGREY)

ws['B32'] = 'FCFF margin'
for i in range(8):
    col = get_column_letter(3 + i)
    c = ws.cell(row=32, column=3 + i, value=f'={col}31/{col}9')
    style_formula(c, fmt='0.0%')


# ============================================================
# SHEET 4: DCF
# ============================================================
ws = wb.create_sheet('4. DCF')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 32
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 16

ws['B2'] = 'DISCOUNTED CASH FLOW VALUATION'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

forecast_years = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
for i, y in enumerate(forecast_years):
    style_header(ws.cell(row=4, column=3 + i, value=y))

ws['B5'] = 'Year #'
for i in range(5):
    ws.cell(row=5, column=3 + i, value=i + 1)
    style_input(ws.cell(row=5, column=3 + i), fmt='0')

ws['B6'] = 'Discount period (mid-year)'
for i in range(5):
    c = ws.cell(row=6, column=3 + i, value=i + 0.5)
    style_input(c, fmt='0.0')

ws['B7'] = 'FCFF (link from IS)'
src_cols = ['F', 'G', 'H', 'I', 'J']
for i, sc in enumerate(src_cols):
    c = ws.cell(row=7, column=3 + i,
                value=f"='3. Income Statement'!{sc}31")
    style_link(c)

ws['B8'] = 'WACC'
ws['C8'] = "='5. WACC'!C12"
style_link(ws['C8'], fmt='0.00%')
ws['C8'].fill = PatternFill('solid', start_color=YELLOW)
for i in range(1, 5):
    c = ws.cell(row=8, column=3 + i, value='=$C$8')
    style_formula(c, fmt='0.00%')

ws['B9'] = 'Discount factor'
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=9, column=3 + i, value=f'=1/(1+{col}8)^{col}6')
    style_formula(c, fmt='0.000')

ws['B10'] = 'PV of FCFF'
ws['B10'].font = Font(name='Arial', bold=True)
for i in range(5):
    col = get_column_letter(3 + i)
    c = ws.cell(row=10, column=3 + i, value=f'={col}7*{col}9')
    style_formula(c); c.font = Font(name='Arial', bold=True)

# Terminal value
ws['B12'] = 'TERMINAL VALUE (Gordon growth)'
ws['B12'].font = Font(name='Arial', bold=True, size=11, color=NYU_VIOLET)
ws['B13'] = 'Terminal growth rate'
ws['C13'] = 0.030
style_input(ws['C13'], fmt='0.00%')
ws['C13'].fill = PatternFill('solid', start_color=YELLOW)

ws['B14'] = 'Terminal FCFF'
ws['C14'] = '=G7*(1+C13)'
style_formula(ws['C14'])

ws['B15'] = 'Terminal value (undiscounted)'
ws['C15'] = '=C14/(C8-C13)'
style_formula(ws['C15'])

ws['B16'] = 'PV of terminal value'
ws['C16'] = '=C15*G9'
style_formula(ws['C16'])
ws['C16'].font = Font(name='Arial', bold=True)

# Summary
ws['B18'] = 'VALUATION SUMMARY'
ws['B18'].font = Font(name='Arial', bold=True, size=11, color=NYU_VIOLET)

ws['B19'] = 'Sum of PV(FCFF)'
ws['C19'] = '=SUM(C10:G10)'
style_formula(ws['C19'])

ws['B20'] = 'PV of terminal value'
ws['C20'] = '=C16'
style_formula(ws['C20'])

ws['B21'] = 'Enterprise value'
ws['C21'] = '=C19+C20'
style_formula(ws['C21']); ws['C21'].font = Font(name='Arial', bold=True)
ws['C21'].fill = PatternFill('solid', start_color=NYU_LGREY)

ws['B22'] = 'Less: Net debt (FY25)'
ws['C22'] = 48958
style_input(ws['C22'])

ws['B23'] = 'Equity value'
ws['C23'] = '=C21-C22'
style_formula(ws['C23']); ws['C23'].font = Font(name='Arial', bold=True)

ws['B24'] = 'Diluted shares (mn)'
ws['C24'] = 4969
style_input(ws['C24'])

ws['B25'] = 'Implied price per share'
ws['C25'] = '=C23/C24'
style_formula(ws['C25'], fmt='$#,##0.00')
ws['C25'].font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
ws['C25'].fill = PatternFill('solid', start_color=NYU_VIOLET)

ws['B26'] = 'Spot price'
ws['C26'] = 422.76
style_input(ws['C26'], fmt='$#,##0.00')

ws['B27'] = 'Implied upside / (downside)'
ws['C27'] = '=C25/C26-1'
style_formula(ws['C27'], fmt='0.0%')

ws['B28'] = 'TV % of EV'
ws['C28'] = '=C20/C21'
style_formula(ws['C28'], fmt='0.0%')


# ============================================================
# SHEET 5: WACC
# ============================================================
ws = wb.create_sheet('5. WACC')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 30

ws['B2'] = 'WEIGHTED AVERAGE COST OF CAPITAL'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

inputs = [
    ('Risk-free rate (10Y UST)', 0.0431, '0.00%', '10Y UST close 24-Apr-26 (Yahoo)'),
    ('Equity risk premium',      0.050,  '0.00%', 'Damodaran 2025 implied ERP'),
    ('Levered beta',             1.20,   '0.00',  'Bloomberg 2Y weekly'),
    ('Pre-tax cost of debt',     0.050,  '0.00%', 'Senior notes blended yield'),
    ('Effective tax rate',       0.18,   '0.0%',  'AVGO 10-K FY25'),
    ('Target debt weight',       0.15,   '0.0%',  'Mgmt long-term capital structure'),
]
for i, (label, val, fmt, src) in enumerate(inputs, start=4):
    ws.cell(row=i, column=2, value=label)
    c = ws.cell(row=i, column=3, value=val)
    style_input(c, fmt=fmt)
    c.fill = PatternFill('solid', start_color=YELLOW)
    ws.cell(row=i, column=4, value=src)
    ws.cell(row=i, column=4).font = Font(name='Arial', italic=True, size=9, color='666666')

ws['B11'] = 'Cost of equity (CAPM)'
ws['C11'] = '=C4+C6*C5'
style_formula(ws['C11'], fmt='0.00%')
ws['C11'].font = Font(name='Arial', bold=True)

ws['B12'] = 'WACC'
ws['B12'].font = Font(name='Arial', bold=True)
ws['C12'] = '=(1-C9)*C11+C9*C7*(1-C8)'
style_formula(ws['C12'], fmt='0.00%')
ws['C12'].font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
ws['C12'].fill = PatternFill('solid', start_color=NYU_VIOLET)


# ============================================================
# SHEET 6: Sensitivity
# ============================================================
ws = wb.create_sheet('6. Sensitivity')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 18
for col in 'CDEFGHI':
    ws.column_dimensions[col].width = 11

ws['B2'] = 'DCF SENSITIVITY — IMPLIED PRICE BY WACC × TERMINAL GROWTH'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

ws['B4'] = 'WACC \\ g'
style_header(ws['B4'])

# Read sensitivity from output
import pandas as pd
sens_df = pd.read_csv(f'{OUT_DIR}/sensitivity.csv', index_col=0)
for j, gcol in enumerate(sens_df.columns):
    style_header(ws.cell(row=4, column=3 + j, value=gcol))
for i, w in enumerate(sens_df.index):
    style_header(ws.cell(row=5 + i, column=2, value=w))
    for j, gcol in enumerate(sens_df.columns):
        v = float(sens_df.loc[w, gcol])
        c = ws.cell(row=5 + i, column=3 + j, value=v)
        style_formula(c, fmt='$#,##0')

# Color scale
last_row = 5 + len(sens_df) - 1
last_col = get_column_letter(3 + len(sens_df.columns) - 1)
rng = f'C5:{last_col}{last_row}'
rule = ColorScaleRule(start_type='min', start_color='F8696B',
                      mid_type='percentile', mid_value=50, mid_color='FFEB84',
                      end_type='max', end_color='63BE7B')
ws.conditional_formatting.add(rng, rule)


# ============================================================
# SHEET 7: Comps
# ============================================================
ws = wb.create_sheet('7. Comps')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 28
for col in 'DEFGHIJKL':
    ws.column_dimensions[col].width = 12

ws['B2'] = 'TRADING COMPARABLES'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

peers_df = pd.read_csv(f'{DATA_DIR}/peer_metrics.csv')
hdrs = ['Ticker', 'Name', 'Price', 'MktCap_$B', 'EV_$B', 'Fwd P/E',
        'EV/Rev', 'EV/EBITDA', 'GM %', 'Op M %', 'Rev Gr %', 'Beta']
for j, h in enumerate(hdrs):
    style_header(ws.cell(row=4, column=2 + j, value=h))

for i, row in peers_df.iterrows():
    ws.cell(row=5 + i, column=2, value=row['Ticker'])
    ws.cell(row=5 + i, column=3, value=row['Name'])
    ws.cell(row=5 + i, column=4, value=row['Price']);          style_input(ws.cell(row=5 + i, column=4), fmt='$#,##0.00')
    ws.cell(row=5 + i, column=5, value=row['MarketCap_B']);    style_input(ws.cell(row=5 + i, column=5), fmt='$#,##0')
    ws.cell(row=5 + i, column=6, value=row['EV_B']);           style_input(ws.cell(row=5 + i, column=6), fmt='$#,##0')
    ws.cell(row=5 + i, column=7, value=row['ForwardPE']);      style_input(ws.cell(row=5 + i, column=7), fmt='0.0"x"')
    ws.cell(row=5 + i, column=8, value=row['EV_to_Revenue']);  style_input(ws.cell(row=5 + i, column=8), fmt='0.0"x"')
    ws.cell(row=5 + i, column=9, value=row['EV_to_EBITDA']);   style_input(ws.cell(row=5 + i, column=9), fmt='0.0"x"')
    ws.cell(row=5 + i, column=10, value=row['GrossMargin']);   style_input(ws.cell(row=5 + i, column=10), fmt='0.0%')
    ws.cell(row=5 + i, column=11, value=row['OpMargin']);      style_input(ws.cell(row=5 + i, column=11), fmt='0.0%')
    ws.cell(row=5 + i, column=12, value=row['RevenueGrowth']); style_input(ws.cell(row=5 + i, column=12), fmt='0.0%')
    ws.cell(row=5 + i, column=13, value=row['Beta']);          style_input(ws.cell(row=5 + i, column=13), fmt='0.00')
    if row['Ticker'] == 'AVGO':
        for col in range(2, 14):
            ws.cell(row=5 + i, column=col).fill = PatternFill('solid', start_color=NYU_LGREY)
            ws.cell(row=5 + i, column=col).font = Font(name='Arial', bold=True, color=BLUE)

last = 5 + len(peers_df) - 1
ws.cell(row=last + 2, column=3, value='Median (peers ex-AVGO)')
ws.cell(row=last + 2, column=3).font = Font(name='Arial', italic=True, bold=True)
for col_letter, ws_col in zip(['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'],
                              range(4, 14)):
    formula = f'=MEDIAN(IF($B$5:$B${last}<>"AVGO",{col_letter}5:{col_letter}{last}))'
    c = ws.cell(row=last + 2, column=ws_col,
                value=f'=MEDIAN({col_letter}5:{col_letter}{last})')
    style_formula(c)
    if ws_col in (7, 8, 9):
        c.number_format = '0.0"x"'
    elif ws_col in (10, 11, 12):
        c.number_format = '0.0%'
    elif ws_col == 13:
        c.number_format = '0.00'
    elif ws_col == 4:
        c.number_format = '$#,##0.00'
    else:
        c.number_format = '#,##0'

# Implied price section
ws['B20'] = 'IMPLIED PRICE FROM PEER MULTIPLES'
ws['B20'].font = Font(name='Arial', bold=True, size=12, color=NYU_VIOLET)
ws['B22'] = 'CY27E EPS estimate ($)'
ws['C22'] = 20.00
style_input(ws['C22'], fmt='$0.00')
ws['C22'].fill = PatternFill('solid', start_color=YELLOW)
ws['B23'] = 'Target Fwd P/E (low)'; ws['C23'] = 18; style_input(ws['C23'], fmt='0.0"x"')
ws['B24'] = 'Target Fwd P/E (mid)'; ws['C24'] = 22; style_input(ws['C24'], fmt='0.0"x"'); ws['C24'].fill = PatternFill('solid', start_color=YELLOW)
ws['B25'] = 'Target Fwd P/E (high)'; ws['C25'] = 26; style_input(ws['C25'], fmt='0.0"x"')
ws['B26'] = 'Implied price (low)';  ws['C26'] = '=C22*C23'; style_formula(ws['C26'], fmt='$#,##0')
ws['B27'] = 'Implied price (mid)';  ws['C27'] = '=C22*C24'; style_formula(ws['C27'], fmt='$#,##0')
ws['B28'] = 'Implied price (high)'; ws['C28'] = '=C22*C25'; style_formula(ws['C28'], fmt='$#,##0')


# ============================================================
# SHEET 8: SOTP
# ============================================================
ws = wb.create_sheet('8. SOTP')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 30
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 16

ws['B2'] = 'SUM-OF-THE-PARTS (FY27E)'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

for j, h in enumerate(['Segment', 'FY27E Rev ($mn)', 'EBITDA Margin', 'FY27E EBITDA',
                       'EV Mult Low', 'EV Mult Mid', 'EV Mult High']):
    style_header(ws.cell(row=4, column=2 + j, value=h))

# Semi
ws['B5'] = 'Semiconductor Solutions'
ws['C5'] = "='2. Assumptions'!G6+'2. Assumptions'!G7"
style_link(ws['C5'])
ws['D5'] = 0.62; style_input(ws['D5'], fmt='0.0%')
ws['D5'].fill = PatternFill('solid', start_color=YELLOW)
ws['E5'] = '=C5*D5'; style_formula(ws['E5'])
ws['F5'] = 16; style_input(ws['F5'], fmt='0.0"x"')
ws['G5'] = 20; style_input(ws['G5'], fmt='0.0"x"'); ws['G5'].fill = PatternFill('solid', start_color=YELLOW)
ws['H5'] = 24; style_input(ws['H5'], fmt='0.0"x"')

# Software
ws['B6'] = 'Infrastructure Software'
ws['C6'] = "='2. Assumptions'!G8"; style_link(ws['C6'])
ws['D6'] = 0.77; style_input(ws['D6'], fmt='0.0%')
ws['D6'].fill = PatternFill('solid', start_color=YELLOW)
ws['E6'] = '=C6*D6'; style_formula(ws['E6'])
ws['F6'] = 14; style_input(ws['F6'], fmt='0.0"x"')
ws['G6'] = 17; style_input(ws['G6'], fmt='0.0"x"'); ws['G6'].fill = PatternFill('solid', start_color=YELLOW)
ws['H6'] = 20; style_input(ws['H6'], fmt='0.0"x"')

# EV by scenario
for j, mlt_col in enumerate(['F', 'G', 'H']):
    ws.cell(row=8, column=6 + j,
            value=f'=E5*{mlt_col}5+E6*{mlt_col}6')
    style_formula(ws.cell(row=8, column=6 + j))
    ws.cell(row=8, column=6 + j).font = Font(name='Arial', bold=True)
ws['B8'] = 'Total EV ($mn)'
ws['B8'].font = Font(name='Arial', bold=True)

ws['B9'] = 'Less: Net debt'
ws['F9'] = 48958; style_input(ws['F9'])
ws['G9'] = 48958; style_input(ws['G9'])
ws['H9'] = 48958; style_input(ws['H9'])

ws['B10'] = 'Equity value'
for j, c in enumerate(['F', 'G', 'H']):
    ws.cell(row=10, column=6 + j, value=f'={c}8-{c}9'); style_formula(ws.cell(row=10, column=6 + j))

ws['B11'] = 'Diluted shares (mn)'
for c in ['F', 'G', 'H']:
    ws[f'{c}11'] = 4969; style_input(ws[f'{c}11'])

ws['B12'] = 'Implied price per share'
ws['B12'].font = Font(name='Arial', bold=True)
for j, c in enumerate(['F', 'G', 'H']):
    ws.cell(row=12, column=6 + j, value=f'={c}10/{c}11')
    style_formula(ws.cell(row=12, column=6 + j), fmt='$#,##0')
    ws.cell(row=12, column=6 + j).font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws.cell(row=12, column=6 + j).fill = PatternFill('solid', start_color=NYU_VIOLET)


# ============================================================
# SHEET 9: Monte Carlo summary
# ============================================================
ws = wb.create_sheet('9. Monte Carlo')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 36
ws.column_dimensions['C'].width = 16

ws['B2'] = 'MONTE CARLO SIMULATION SUMMARY'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)
ws['B3'] = '10,000 paths • Drivers in scripts/02_valuation.py'
ws['B3'].font = Font(name='Arial', italic=True, size=9, color='666666')

import json

with open(f'{OUT_DIR}/summary.json') as f:
    summary = json.load(f)

stats = [
    ('Mean intrinsic price',          summary['mc_mean']),
    ('Median',                        summary['mc_median']),
    ('25th percentile (P25)',         summary['mc_p25']),
    ('75th percentile (P75)',         summary['mc_p75']),
    ('Probability(intrinsic > spot)', summary['mc_prob_above_spot']),
]
for i, (label, val) in enumerate(stats, start=5):
    ws.cell(row=i, column=2, value=label)
    c = ws.cell(row=i, column=3, value=val)
    if 'Probability' in label:
        style_input(c, fmt='0.0%')
    else:
        style_input(c, fmt='$#,##0.00')

ws['B12'] = 'Driver distribution (assumed Normal)'
ws['B12'].font = Font(name='Arial', bold=True, color=NYU_DARK)
drivers = [
    ('FY27 AI revenue ($B)', '~ N(125, 25)'),
    ('AI growth FY27→FY30 CAGR', '~ N(30%, 10%)'),
    ('Terminal growth rate', '~ N(3.0%, 0.5%)'),
    ('WACC', '~ N(base WACC, 0.5%)'),
    ('Steady-state EBITDA margin', '~ N(68.5%, 2.5%)'),
    ('Capex / Sales', '~ N(1.1%, 0.3%)'),
    ('Non-AI semi growth', '~ N(4%, 2%)'),
    ('Software growth', '~ N(7%, 2.5%)'),
]
for i, (drv, dist) in enumerate(drivers, start=13):
    ws.cell(row=i, column=2, value='  ' + drv)
    ws.cell(row=i, column=3, value=dist)
    ws.cell(row=i, column=3).font = Font(name='Arial', italic=True, color='333333')


# ============================================================
# SHEET 10: Football Field
# ============================================================
ws = wb.create_sheet('10. Football Field')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 32
for col in 'CDEFG':
    ws.column_dimensions[col].width = 14

ws['B2'] = 'VALUATION SYNTHESIS — FOOTBALL FIELD'
ws['B2'].font = Font(name='Arial', bold=True, size=14, color=NYU_VIOLET)

for j, h in enumerate(['Method', 'Low', 'Mid', 'High', 'Implied Upside (Mid)']):
    style_header(ws.cell(row=4, column=2 + j, value=h))

football = pd.read_csv(f'{OUT_DIR}/football_field.csv')
spot = 422.76
for i, row in football.iterrows():
    ws.cell(row=5 + i, column=2, value=row['Method'])
    ws.cell(row=5 + i, column=3, value=row['Low']);  style_input(ws.cell(row=5 + i, column=3), fmt='$#,##0')
    ws.cell(row=5 + i, column=4, value=row['Mid']);  style_input(ws.cell(row=5 + i, column=4), fmt='$#,##0')
    ws.cell(row=5 + i, column=5, value=row['High']); style_input(ws.cell(row=5 + i, column=5), fmt='$#,##0')
    c = ws.cell(row=5 + i, column=6, value=f'=D{5 + i}/{spot}-1')
    style_formula(c, fmt='0.0%')

last = 5 + len(football) - 1
ws.cell(row=last + 2, column=2, value='Average').font = Font(name='Arial', bold=True)
for col_letter, col_idx in zip(['C', 'D', 'E'], [3, 4, 5]):
    c = ws.cell(row=last + 2, column=col_idx,
                value=f'=AVERAGE({col_letter}5:{col_letter}{last})')
    style_formula(c, fmt='$#,##0'); c.font = Font(name='Arial', bold=True)
    c.fill = PatternFill('solid', start_color=NYU_LGREY)

ws.cell(row=last + 4, column=2, value='Spot price')
ws.cell(row=last + 4, column=3, value=spot); style_input(ws.cell(row=last + 4, column=3), fmt='$#,##0.00')

ws.cell(row=last + 5, column=2, value='Recommendation').font = Font(name='Arial', bold=True)
ws.cell(row=last + 5, column=3, value='BUY')
ws.cell(row=last + 5, column=3).font = Font(name='Arial', bold=True, color='FFFFFF')
ws.cell(row=last + 5, column=3).fill = PatternFill('solid', start_color=ACCENT_GOLD)


wb.save(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH}")
